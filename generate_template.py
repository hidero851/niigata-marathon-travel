"""
大会情報入力テンプレート（Excel）生成スクリプト
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===================== スタイル定義 =====================
NAVY   = "1A2E5A"
ORANGE = "F97316"
BLUE   = "DBEAFE"
GREEN  = "DCFCE7"
YELLOW = "FEF9C3"
GRAY   = "F1F5F9"
WHITE  = "FFFFFF"
RED    = "FEE2E2"

def header_style(cell, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size, name="Meiryo")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def sub_header(cell, bg=GRAY):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, size=10, name="Meiryo")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def note_cell(cell, bg=YELLOW):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(size=9, color="78350F", name="Meiryo")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def input_cell(cell, bg=WHITE):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(size=10, name="Meiryo")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def required_label(cell):
    cell.fill = PatternFill("solid", fgColor="FEE2E2")
    cell.font = Font(bold=True, size=10, name="Meiryo")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def optional_label(cell):
    cell.fill = PatternFill("solid", fgColor=GRAY)
    cell.font = Font(size=10, name="Meiryo", color="475569")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def title_row(ws, row, text, note=""):
    ws.row_dimensions[row].height = 36
    c = ws.cell(row=row, column=1, value=text)
    header_style(c, bg=NAVY, size=13)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    if note:
        ws.row_dimensions[row+1].height = 28
        n = ws.cell(row=row+1, column=1, value=note)
        note_cell(n)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=6)
    return row + (2 if note else 1)

def field_row(ws, row, label, example, required=True, note=""):
    ws.row_dimensions[row].height = 24
    lbl = ws.cell(row=row, column=1, value=("★ " if required else "　") + label)
    if required:
        required_label(lbl)
    else:
        optional_label(lbl)
    ex = ws.cell(row=row, column=2, value=f"例）{example}" if example else "")
    note_cell(ex, bg=YELLOW if example else WHITE)
    inp = ws.cell(row=row, column=3, value="")
    input_cell(inp)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    if note:
        ws.row_dimensions[row+1].height = 20
        n = ws.cell(row=row+1, column=1, value=f"  ↑ {note}")
        n.fill = PatternFill("solid", fgColor="F8FAFC")
        n.font = Font(size=8, color="94A3B8", name="Meiryo")
        n.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=6)
        return row + 2
    return row + 1

# ===================== シート1: 大会基本情報 =====================
ws1 = wb.active
ws1.title = "①大会基本情報"
set_col_widths(ws1, {"A": 28, "B": 30, "C": 20, "D": 20, "E": 20, "F": 20})

r = 1
# タイトル
ws1.row_dimensions[r].height = 48
t = ws1.cell(row=r, column=1, value="新潟マラソントラベル　大会情報入力テンプレート")
header_style(t, bg=NAVY, size=15)
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
ws1.row_dimensions[r].height = 24
n = ws1.cell(row=r, column=1, value="★マーク：必須入力　　記入後、担当者（hidero）へ送付してください")
note_cell(n, bg="FFF7ED")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 2

r = title_row(ws1, r, "【セクション1】大会の基本情報", "大会の概要・開催概要を入力してください")
r = field_row(ws1, r, "大会名", "高田城ロードレース")
r = field_row(ws1, r, "開催地（市区町村）", "上越市")
r = field_row(ws1, r, "都道府県", "新潟県")
r = field_row(ws1, r, "開催日（表示用）", "2026年6月7日（日）予定")
r = field_row(ws1, r, "開催日（ISO形式）", "2026-06-07", note="YYYY-MM-DD形式で入力（前泊推奨日の自動計算に使用）")
r = field_row(ws1, r, "開催月（数字）", "6", note="4〜11 の数字のみ")
r = field_row(ws1, r, "種目・距離", "ハーフ / 10km / 5km", note="スラッシュ区切りで入力")
r = field_row(ws1, r, "キャッチコピー", "城下町を走り、上越の春を楽しむ。")
r = field_row(ws1, r, "大会の説明文（短め）", "高田城址公園を舞台にしたロードレース。ファミリーにも人気。")
r += 1

r = title_row(ws1, r, "【セクション2】参加費・定員・制限")
r = field_row(ws1, r, "参加費", "フルマラソン: 8,000円　ハーフ: 5,000円", required=False)
r = field_row(ws1, r, "定員", "5,000人", required=False)
r = field_row(ws1, r, "制限時間", "3時間30分", required=False)
r = field_row(ws1, r, "スタート地点", "高田城址公園陸上競技場", required=False)
r = field_row(ws1, r, "ゴール地点", "高田城址公園陸上競技場", required=False)
r = field_row(ws1, r, "会場（施設名）", "高田城址公園", required=False)
r = field_row(ws1, r, "申込期間", "2026年1月10日〜3月31日", required=False)
r = field_row(ws1, r, "主催者", "上越市体育協会", required=False)
r = field_row(ws1, r, "アクセス", "えちごトキめき鉄道「高田駅」から徒歩15分", required=False)
r = field_row(ws1, r, "備考・注意事項", "雨天決行・荒天中止", required=False)
r += 1

r = title_row(ws1, r, "【セクション3】URL・タグ情報")
r = field_row(ws1, r, "エントリーURL（RUNNET等）", "https://runnet.jp/entry/xxxxx", required=False)
r = field_row(ws1, r, "公式サイトURL", "https://www.city.joetsu.niigata.jp/", required=False)
r = field_row(ws1, r, "ヒーロー画像URL", "https://example.com/hero.jpg", required=False, note="大会の雰囲気が伝わる横長画像のURLを記入")
r = field_row(ws1, r, "タグ（カンマ区切り）", "城下町, 家族向け, 初夏", note="温泉 / グルメ / 絶景 / 佐渡 / 桜 / 城下町 / 地酒 / 米どころ 等から選択")

# ===================== シート2: 楽しめること =====================
ws2 = wb.create_sheet("②楽しめること")
set_col_widths(ws2, {"A": 6, "B": 24, "C": 40, "D": 35, "E": 6})
ws2.row_dimensions[1].height = 48
t = ws2.cell(row=1, column=1, value="【このレースで楽しめること】入力シート")
header_style(t, bg=NAVY, size=13)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

ws2.row_dimensions[2].height = 24
n = ws2.cell(row=2, column=1, value="最大4件まで入力できます。大会の魅力・見どころを具体的に書いてください。")
note_cell(n, bg="FFF7ED")
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

ws2.row_dimensions[3].height = 28
for col, val, bg in [(1,"No.",(GRAY)),(2,"タイトル（15文字以内）",(NAVY)),(3,"説明文（50文字程度）",(NAVY)),(4,"画像URL（任意）",(NAVY))]:
    c = ws2.cell(row=3, column=col, value=val)
    header_style(c, bg=bg if col > 1 else GRAY, fg=WHITE if col > 1 else "1E293B", size=10)

for i, (title, desc) in enumerate([
    ("高田城址公園を走る", "国内屈指の桜の名所を駆け抜ける、春ならではの特別なルート。"),
    ("城下町の風情", "歴史情緒あふれる高田の街並みとともに走る、唯一無二の体験。"),
    ("上越グルメを楽しむ", "笹団子・地酒・日本海の海鮮でゴール後の食を満喫できます。"),
    ("家族みんなで参加", "5km部門もあり、ファミリー全員で楽しめる大会です。"),
], start=4):
    ws2.row_dimensions[i].height = 60
    ws2.cell(row=i, column=1, value=i-3).alignment = Alignment(horizontal="center", vertical="center")
    for col, val in [(2, f"例）{title}"),(3, f"例）{desc}"),(4,"")]:
        c = ws2.cell(row=i, column=col, value=val)
        if "例）" in str(val):
            note_cell(c, bg=YELLOW)
        else:
            input_cell(c)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 実際の入力行（8〜11行）
for i in range(8, 12):
    ws2.row_dimensions[i].height = 60
    ws2.cell(row=i, column=1, value=i-7).alignment = Alignment(horizontal="center", vertical="center")
    for col in [2, 3, 4]:
        c = ws2.cell(row=i, column=col, value="")
        input_cell(c)

ws2.row_dimensions[7].height = 24
sep = ws2.cell(row=7, column=1, value="↓ こちらに入力してください")
sep.fill = PatternFill("solid", fgColor="DBEAFE")
sep.font = Font(bold=True, size=10, color="1D4ED8", name="Meiryo")
sep.alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells(start_row=7, start_column=1, end_row=7, end_column=5)

# ===================== シート3: 食・特産・お土産 =====================
ws3 = wb.create_sheet("③食・特産・お土産")
set_col_widths(ws3, {"A": 6, "B": 22, "C": 16, "D": 30, "E": 30, "F": 20, "G": 30})
ws3.row_dimensions[1].height = 48
t = ws3.cell(row=1, column=1, value="【食・特産・お土産】入力シート")
header_style(t, bg=NAVY, size=13)
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

ws3.row_dimensions[2].height = 24
n = ws3.cell(row=2, column=1, value="大会開催エリアの特産品・お土産・グルメを入力してください。複数件入力可。")
note_cell(n, bg="FFF7ED")
ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

headers = ["No.", "商品名・店名", "エリア", "短い説明（30文字）", "詳細説明（100文字程度）", "購入場所・販売店", "外部URL（任意）"]
ws3.row_dimensions[3].height = 28
for col, h in enumerate(headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    header_style(c, bg=NAVY if col > 1 else GRAY, fg=WHITE if col > 1 else "1E293B", size=10)

examples = [
    ("笹団子", "上越市", "上越名物の草餅系和菓子", "よもぎを練り込んだ餅にあんこを包み、笹の葉で包んだ上越を代表する郷土菓子。", "高田駅構内・市内和菓子店", ""),
    ("上越の地酒", "上越市", "雪国育ちの辛口地酒", "豪雪地帯の良質な水と米で醸した新潟らしいキレのある日本酒。蔵元直売所でも購入可。", "市内酒蔵・道の駅", ""),
    ("のどぐろ干物", "上越市・直江津", "日本海で獲れた高級魚", "「白身のトロ」と呼ばれるのどぐろを使った干物。旨味が凝縮されたお土産の定番。", "直江津港周辺・鮮魚店", ""),
]
for i, ex in enumerate(examples, 4):
    ws3.row_dimensions[i].height = 55
    ws3.cell(row=i, column=1, value=f"例{i-3}").alignment = Alignment(horizontal="center", vertical="center")
    for col, val in enumerate(ex, 2):
        c = ws3.cell(row=i, column=col, value=f"例）{val}" if val else "")
        note_cell(c, bg=YELLOW)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws3.row_dimensions[7].height = 24
sep = ws3.cell(row=7, column=1, value="↓ こちらに入力してください（何件でも追加可）")
sep.fill = PatternFill("solid", fgColor="DBEAFE")
sep.font = Font(bold=True, size=10, color="1D4ED8", name="Meiryo")
sep.alignment = Alignment(horizontal="left", vertical="center")
ws3.merge_cells(start_row=7, start_column=1, end_row=7, end_column=7)

for i in range(8, 18):
    ws3.row_dimensions[i].height = 55
    ws3.cell(row=i, column=1, value=i-7).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, 8):
        c = ws3.cell(row=i, column=col, value="")
        input_cell(c)

# ===================== シート4: 宿泊エリア =====================
ws4 = wb.create_sheet("④宿泊エリア")
set_col_widths(ws4, {"A": 6, "B": 22, "C": 20, "D": 20, "E": 40, "F": 20})
ws4.row_dimensions[1].height = 48
t = ws4.cell(row=1, column=1, value="【宿泊エリア情報】入力シート")
header_style(t, bg=NAVY, size=13)
ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

ws4.row_dimensions[2].height = 24
n = ws4.cell(row=2, column=1, value="ランナー向けに推奨する宿泊エリアを3つ入力してください（🥇当日ラク重視 / 🥈前泊バランス / 🥉観光も楽しむ）")
note_cell(n, bg="FFF7ED")
ws4.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

headers4 = ["順位", "エリア名", "会場からの距離感", "価格帯目安", "エリアの説明（80文字程度）", "このエリアで楽しめること"]
ws4.row_dimensions[3].height = 28
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=3, column=col, value=h)
    header_style(c, bg=NAVY if col > 1 else GRAY, fg=WHITE if col > 1 else "1E293B", size=10)

stays = [
    ("🥇 当日ラク重視", "高田駅周辺", "会場まで徒歩〜車10分", "8,000〜15,000円", "飲食店が多く前泊後の夕食にも便利。大会当日も会場へのアクセスが良好なエリア。", "城下町散策 / 居酒屋・地酒"),
    ("🥈 前泊バランス", "直江津・上越妙高周辺", "会場まで車・電車20〜30分", "7,000〜12,000円", "新幹線や日本海観光と組み合わせやすい。飲食店も充実し前泊向き。", "日本海の海鮮 / 温泉"),
    ("🥉 観光も楽しむ", "妙高・糸魚川方面", "会場まで車40〜60分", "10,000〜20,000円", "大会＋温泉・絶景観光をセットで楽しみたい人向け。家族旅行との相性が良い。", "温泉 / 妙高の絶景 / 糸魚川ジオパーク"),
]
for i, s in enumerate(stays, 4):
    ws4.row_dimensions[i].height = 65
    ws4.cell(row=i, column=1, value=s[0]).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, val in enumerate(s[1:], 2):
        c = ws4.cell(row=i, column=col, value=f"例）{val}")
        note_cell(c, bg=YELLOW)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws4.row_dimensions[7].height = 24
sep = ws4.cell(row=7, column=1, value="↓ こちらに入力してください")
sep.fill = PatternFill("solid", fgColor="DBEAFE")
sep.font = Font(bold=True, size=10, color="1D4ED8", name="Meiryo")
sep.alignment = Alignment(horizontal="left", vertical="center")
ws4.merge_cells(start_row=7, start_column=1, end_row=7, end_column=6)

for i, rank in enumerate(["🥇 当日ラク重視", "🥈 前泊バランス", "🥉 観光も楽しむ"], 8):
    ws4.row_dimensions[i].height = 65
    ws4.cell(row=i, column=1, value=rank).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(2, 7):
        c = ws4.cell(row=i, column=col, value="")
        input_cell(c)

# ===================== シート5: 参加プラン =====================
ws5 = wb.create_sheet("⑤おすすめ参加プラン")
set_col_widths(ws5, {"A": 22, "B": 5, "C": 50, "D": 5})
ws5.row_dimensions[1].height = 48
t = ws5.cell(row=1, column=1, value="【おすすめ参加プラン（モデルコース）】入力シート")
header_style(t, bg=NAVY, size=13)
ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

ws5.row_dimensions[2].height = 24
n = ws5.cell(row=2, column=1, value="前泊プランと当日日帰りプランの2パターンを想定しています。1〜2プラン入力してください。")
note_cell(n, bg="FFF7ED")
ws5.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

plans = [
    ("前泊＋観光プラン", [
        "【前日】現地入り・高田の宿に宿泊",
        "【前日夜】城下町の居酒屋で地酒・海鮮を楽しむ",
        "【当日朝】会場へ移動・ウォームアップ",
        "【レース】ハーフマラソン完走！",
        "【ゴール後】高田城址公園を散策",
        "【午後】笹団子・お土産を購入して帰路へ",
    ]),
    ("当日日帰りプラン", [
        "【朝早め】新幹線・車で上越市へ",
        "【午前】会場到着・受付・アップ",
        "【レース】10km / ハーフ完走！",
        "【ゴール後】直江津港で海鮮ランチ",
        "【午後】お土産購入後、帰路へ",
    ]),
]
r5 = 4
for plan_name, steps in plans:
    ws5.row_dimensions[r5].height = 32
    h = ws5.cell(row=r5, column=1, value=f"プラン名：{plan_name}")
    header_style(h, bg=ORANGE, size=11)
    ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=4)
    r5 += 1
    for j, step in enumerate(steps, 1):
        ws5.row_dimensions[r5].height = 24
        n = ws5.cell(row=r5, column=1, value=f"例）ステップ{j}: {step}")
        note_cell(n, bg=YELLOW)
        ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=4)
        r5 += 1
    # 入力行
    ws5.row_dimensions[r5].height = 20
    sep2 = ws5.cell(row=r5, column=1, value="↓ 実際に入力してください")
    sep2.fill = PatternFill("solid", fgColor="DBEAFE")
    sep2.font = Font(bold=True, size=9, color="1D4ED8", name="Meiryo")
    ws5.merge_cells(start_row=r5, start_column=1, end_row=r5, end_column=4)
    r5 += 1
    in_name = ws5.cell(row=r5, column=1, value="プラン名：")
    sub_header(in_name)
    inp = ws5.cell(row=r5, column=2, value="")
    input_cell(inp)
    ws5.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=4)
    r5 += 1
    for j in range(1, 7):
        ws5.row_dimensions[r5].height = 28
        lbl = ws5.cell(row=r5, column=1, value=f"ステップ{j}")
        sub_header(lbl)
        inp = ws5.cell(row=r5, column=2, value="")
        input_cell(inp)
        ws5.merge_cells(start_row=r5, start_column=2, end_row=r5, end_column=4)
        r5 += 1
    r5 += 1

# ===================== シート6: 記入ガイド =====================
ws6 = wb.create_sheet("⑥記入ガイド")
set_col_widths(ws6, {"A": 28, "B": 50})
ws6.row_dimensions[1].height = 48
t = ws6.cell(row=1, column=1, value="記入ガイド・よくある質問")
header_style(t, bg=NAVY, size=14)
ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

guides = [
    ("★マーク（必須）とは？", "★マークがついている項目はサイト表示に必要な必須項目です。"),
    ("画像URLとは？", "画像ファイルをDropbox・Googleドライブ等に保存し、共有リンクのURLを貼ってください。"),
    ("開催日のISO形式とは？", "YYYY-MM-DD形式（例: 2026-06-07）で入力してください。この形式が前泊推奨日の自動計算に使われます。"),
    ("タグはどこから選ぶ？", "温泉 / グルメ / 絶景 / 佐渡 / 桜 / 城下町 / 地酒 / 米どころ / 家族向け / 自然 / 初心者向け 等から選択。カンマ区切りで複数入力可。"),
    ("分からない項目は？", "「要確認」「未定」と入力してください。空白のままでもOKです。"),
    ("送付先・提出方法は？", "記入済みのExcelファイルをメールまたはチャットで担当者へ送付してください。"),
    ("修正・追加がある場合は？", "修正後のファイルを再度送付してください。インポート時に上書きで反映します。"),
]
r6 = 3
for q, a in guides:
    ws6.row_dimensions[r6].height = 28
    qc = ws6.cell(row=r6, column=1, value=q)
    sub_header(qc, bg=BLUE)
    qc.font = Font(bold=True, size=10, name="Meiryo", color="1D4ED8")
    ac = ws6.cell(row=r6, column=2, value=a)
    input_cell(ac)
    ac.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    r6 += 1

# ===================== 保存 =====================
output_path = r"C:\Users\hider\niigata-marathon-travel\大会情報入力テンプレート.xlsx"
wb.save(output_path)
print(f"作成完了: {output_path}")
